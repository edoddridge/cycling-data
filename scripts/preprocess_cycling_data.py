from __future__ import annotations

import argparse
import json
import re
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Iterable

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
RAW_DATA_DIR = ROOT / "data" / "cycling"
OUTPUT_DIR = ROOT / "data" / "processed"

DATASETS = [
    {
        "id": "st",
        "name": "Super Tuesday (Commuter)",
        "source": RAW_DATA_DIR / "ST_all_index PBI 2010-2025 Download.xlsx",
        "output": OUTPUT_DIR / "st.json",
        "categories": [
            {"id": "female", "field": "female", "label": "Pushbike women", "color": "#0077b6"},
            {"id": "male", "field": "male", "label": "Pushbike men", "color": "#48cae4"},
            {"id": "not_known", "field": "not known", "label": "Pushbike unknown", "color": "#90e0ef"},
            {"id": "women_ebike", "field": "women ebike", "label": "E-bike women", "color": "#ef476f"},
            {"id": "men_ebike", "field": "men ebike", "label": "E-bike men", "color": "#f78c6b"},
            {"id": "not_known_ebike", "field": "not known ebike", "label": "E-bike unknown", "color": "#ffcad4"},
            {"id": "women_mm", "field": "women mm", "label": "Micro mobility women", "color": "#7b2cbf"},
            {"id": "men_mm", "field": "men mm", "label": "Micro mobility men", "color": "#9d4edd"},
            {"id": "not_known_mm", "field": "not known mm", "label": "Micro mobility unknown", "color": "#c77dff"},
        ],
    },
    {
        "id": "ss",
        "name": "Super Sunday (Recreation)",
        "source": RAW_DATA_DIR / "SS_all_index PBI 2010-2024 Download.xlsx",
        "output": OUTPUT_DIR / "ss.json",
        "categories": [
            {"id": "bicycle", "field": "bicycle", "label": "Bicycle", "color": "#2a9d8f"},
            {"id": "ebike", "field": "e-bike", "label": "E-bike", "color": "#264653"},
            {"id": "walker", "field": "walker", "label": "Walker", "color": "#e9c46a"},
            {"id": "runner", "field": "runner", "label": "Runner", "color": "#f4a261"},
            {"id": "dog", "field": "dog", "label": "Dog walker", "color": "#e76f51"},
            {"id": "micro", "field": "micro", "label": "Micro mobility", "color": "#6a4c93"},
            {"id": "other", "field": "other", "label": "Other", "color": "#8d99ae"},
        ],
    },
]

TIME_COLUMN_PATTERN = re.compile(r"^\d{2}:\d{2}:\d{2}$")
TIME_IN_TEXT_PATTERN = re.compile(r"(?<!\d)(\d{1,2})[:._-](\d{2})(?::(\d{2}))?(?!\d)")


def as_number(value: Any) -> int | float | None:
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        return value

    text = str(value).strip()
    if not text or text.lower() == "no data":
        return None

    text = text.replace(",", "")
    try:
        numeric = float(text)
    except ValueError:
        return None
    return int(numeric) if numeric.is_integer() else numeric


def parse_date_value(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        return date(1899, 12, 30) + timedelta(days=float(value))

    text = str(value).strip()
    if not text:
        return None

    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def format_date_key(value: date | None) -> str:
    return value.isoformat() if value else "unknown"


def format_date_label(value: date | None) -> str:
    return value.isoformat() if value else "unknown"


def normalize_key(key: Any) -> str:
    return str(key).strip().lower()


def normalize_row(row: dict[str, Any]) -> dict[str, Any]:
    return {normalize_key(key): value for key, value in row.items()}


def extract_time_columns(headers: Iterable[str]) -> list[str]:
    return sorted([header for header in headers if TIME_COLUMN_PATTERN.match(str(header or ""))])


def format_time_label(hour: int, minute: int) -> str:
    return f"{hour:02d}:{minute:02d}"


def extract_time_from_text(value: str) -> str | None:
    match = TIME_IN_TEXT_PATTERN.search(value)
    if not match:
        return None

    hour = int(match.group(1))
    minute = int(match.group(2))
    if hour < 0 or hour > 23 or minute < 0 or minute > 59:
        return None
    return format_time_label(hour, minute)


def tokenize_key(value: str) -> list[str]:
    return re.findall(r"[a-z0-9]+", normalize_key(value))


def resolve_category_hourly_columns(dataset: dict[str, Any], headers: Iterable[str]) -> dict[str, list[tuple[str, str]]]:
    columns_by_category: dict[str, dict[str, str]] = {definition["id"]: {} for definition in dataset["categories"]}

    for header in headers:
        normalized_header = normalize_key(header)
        if not normalized_header:
            continue

        time_label = extract_time_from_text(normalized_header)
        if not time_label:
            continue

        header_without_time = TIME_IN_TEXT_PATTERN.sub(" ", normalized_header)
        header_tokens = set(tokenize_key(header_without_time))
        if not header_tokens:
            continue

        for definition in dataset["categories"]:
            field_tokens = tokenize_key(definition["field"])
            if field_tokens and all(token in header_tokens for token in field_tokens):
                columns_by_category[definition["id"]][time_label] = normalized_header

    resolved: dict[str, list[tuple[str, str]]] = {}
    for category_id, columns in columns_by_category.items():
        if not columns:
            continue
        resolved[category_id] = sorted(columns.items(), key=lambda item: item[0])
    return resolved


def pick_total(normalized_row: dict[str, Any], dataset_id: str) -> int | float | None:
    if dataset_id == "st":
        return as_number(normalized_row.get("total_thisyear")) or as_number(normalized_row.get("total_this_year"))
    return as_number(normalized_row.get("total"))


def extract_road_labels(description: Any) -> list[str]:
    if not description:
        return []
    return [part.strip() for part in str(description).split(",") if part.strip()]


def aggregate_yearly_totals(records: Iterable[dict[str, Any]]) -> dict[int, float]:
    grouped: dict[int, list[float]] = {}
    for record in records:
        year = record.get("year")
        total = record.get("total")
        if year is None or total is None:
            continue
        grouped.setdefault(int(year), []).append(float(total))
    return {year: round(sum(values) / len(values), 2) for year, values in sorted(grouped.items())}


def extract_categories(dataset: dict[str, Any], normalized_row: dict[str, Any]) -> dict[str, int | float]:
    categories: dict[str, int | float] = {}
    for definition in dataset["categories"]:
        value = as_number(normalized_row.get(definition["field"]))
        if value is not None:
            categories[definition["id"]] = value
    return categories


def extract_directional_data(normalized_row: dict[str, Any], legs: int) -> dict[str, list[dict[str, int | float | None]]]:
    leg_totals = []
    for index in range(1, legs + 1):
        leg_totals.append(
            {
                "leg": index,
                "enter": as_number(normalized_row.get(f"leg{index}_enter")),
                "exit": as_number(normalized_row.get(f"leg{index}_exit")),
                "total": as_number(normalized_row.get(f"leg{index}_total")),
            }
        )

    movements = []
    for from_leg in range(1, legs + 1):
        for to_leg in range(1, legs + 1):
            if from_leg == to_leg:
                continue
            value = as_number(normalized_row.get(f"leg{from_leg}-{to_leg}"))
            if value is not None:
                movements.append({"from": from_leg, "to": to_leg, "value": value})

    return {"legTotals": leg_totals, "movements": movements}


def extract_category_hourly_data(normalized_row: dict[str, Any], category_hourly_columns: dict[str, list[tuple[str, str]]]) -> dict[str, list[dict[str, int | float | None]]]:
    category_hourly: dict[str, list[dict[str, int | float | None]]] = {}
    for category_id, columns in category_hourly_columns.items():
        series = []
        has_non_null = False
        for time_label, column_key in columns:
            value = as_number(normalized_row.get(column_key))
            if value is not None:
                has_non_null = True
            series.append({"time": time_label, "value": value})
        if has_non_null:
            category_hourly[category_id] = series
    return category_hourly


def build_site_record(
    dataset: dict[str, str],
    normalized_row: dict[str, Any],
    raw_row: dict[str, Any],
    time_columns: list[str],
    category_hourly_columns: dict[str, list[tuple[str, str]]],
    sites: dict[str, dict[str, Any]],
) -> None:
    site_id = as_number(normalized_row.get("site_id"))
    latitude = as_number(normalized_row.get("latitude"))
    longitude = as_number(normalized_row.get("longitude"))
    if site_id is None or latitude is None or longitude is None:
        return

    site_key = f"{dataset['id']}:{int(site_id)}"
    if site_key not in sites:
        sites[site_key] = {
            "key": site_key,
            "datasetId": dataset["id"],
            "datasetName": dataset["name"],
            "siteId": int(site_id),
            "state": str(normalized_row.get("state") or "").strip(),
            "council": str(normalized_row.get("council") or "").strip(),
            "latitude": latitude,
            "longitude": longitude,
            "description": str(normalized_row.get("description") or "").strip(),
            "legs": int(as_number(normalized_row.get("legs")) or 0),
            "exitLayout": [None] * 6,
            "entryLayout": [None] * 6,
            "roadLabels": extract_road_labels(normalized_row.get("description")),
            "records": [],
        }

    site = sites[site_key]
    for index in range(1, 7):
        exit_value = as_number(normalized_row.get(f"layout_{index}"))
        entry_value = as_number(normalized_row.get(f"layout_{index}_enter"))
        if site["exitLayout"][index - 1] is None and exit_value is not None:
            site["exitLayout"][index - 1] = exit_value
        if site["entryLayout"][index - 1] is None and entry_value is not None:
            site["entryLayout"][index - 1] = entry_value

    parsed_date = parse_date_value(normalized_row.get("count_date"))
    categories = extract_categories(dataset, normalized_row)
    category_hourly = extract_category_hourly_data(normalized_row, category_hourly_columns)
    directional = extract_directional_data(normalized_row, site["legs"])
    site["records"].append(
        {
            "date": format_date_key(parsed_date),
            "year": parsed_date.year if parsed_date else None,
            "total": pick_total(normalized_row, dataset["id"]),
            "categories": categories,
            "categoryHourly": category_hourly,
            "directional": directional,
            "hourly": [
                {"time": time_value[:5], "value": as_number(raw_row.get(time_value))}
                for time_value in time_columns
            ],
        }
    )


def process_workbook(dataset: dict[str, str]) -> dict[str, Any]:
    workbook = openpyxl.load_workbook(dataset["source"], read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    header_cells = next(worksheet.iter_rows(min_row=2, max_row=2, values_only=True))
    headers = [str(cell).strip() if cell is not None else "" for cell in header_cells]
    time_columns = extract_time_columns(headers)
    category_hourly_columns = resolve_category_hourly_columns(dataset, headers)
    sites: dict[str, dict[str, Any]] = {}

    for values in worksheet.iter_rows(min_row=3, values_only=True):
        raw_row = {
            header: values[index] if index < len(values) else None
            for index, header in enumerate(headers)
            if header
        }
        if not raw_row:
            continue
        normalized = normalize_row(raw_row)
        build_site_record(dataset, normalized, raw_row, time_columns, category_hourly_columns, sites)

    workbook.close()

    sorted_sites = sorted(sites.values(), key=lambda site: (site["council"].lower(), site["siteId"]))
    for site in sorted_sites:
        site["records"] = sorted(site["records"], key=lambda record: (record["date"], record["total"] or -1))
        site["yearlyTotals"] = aggregate_yearly_totals(site["records"])

    return {
        "datasetId": dataset["id"],
        "datasetName": dataset["name"],
        "categoryDefinitions": [
            {"id": definition["id"], "label": definition["label"], "color": definition["color"]}
            for definition in dataset["categories"]
        ],
        "siteCount": len(sorted_sites),
        "sites": sorted_sites,
    }


def write_json(payload: dict[str, Any], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, separators=(",", ":")), encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="Preprocess cycling count workbooks into lightweight JSON.")
    parser.parse_args()

    manifest = []
    for dataset in DATASETS:
        payload = process_workbook(dataset)
        write_json(payload, dataset["output"])
        manifest.append(
            {
                "datasetId": payload["datasetId"],
                "datasetName": payload["datasetName"],
                "siteCount": payload["siteCount"],
                "file": str(dataset["output"].relative_to(ROOT)),
            }
        )

    write_json({"datasets": manifest}, OUTPUT_DIR / "index.json")
    print(f"Wrote {len(manifest)} processed dataset files to {OUTPUT_DIR}")


if __name__ == "__main__":
    main()